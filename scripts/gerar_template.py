"""
scripts/gerar_template.py
Gera MODELO_DESPESAS_PREFEITURA.xlsx com:
  - Aba "Instruções" — como preencher
  - Aba "Lançamentos" — onde a prefeitura digita um pagamento por linha
  - Abas de catálogo: Secretarias, Recursos, Categorias, Fornecedores
  - Validação de dados (dropdown) nas colunas-chave
  - Formatação de data e moeda (BRL)
  - Coluna Mês/Ano calculada por fórmula a partir da Data
  - Cabeçalhos congelados e destacados

Uso:
    python scripts/gerar_template.py
"""
from __future__ import annotations
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, NamedStyle,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

OUT_DIR = Path(__file__).resolve().parent.parent / "data" / "templates"
OUT_DIR.mkdir(parents=True, exist_ok=True)
OUT_FILE = OUT_DIR / "MODELO_DESPESAS_PREFEITURA.xlsx"

# -------------------------------------------------------------------
# Catálogos (listas suspensas)
# -------------------------------------------------------------------
SECRETARIAS = [
    "ADMINISTRAÇÃO / RP",
    "ASSISTÊNCIA SOCIAL",
    "CONVÊNIOS",
    "CULTURA E TURISMO",
    "EDUCAÇÃO",
    "GABINETE",
    "OBRAS E SERVIÇOS URBANOS",
    "SAÚDE",
]

# Recursos/Fontes consolidados (sem variações de digitação)
RECURSOS = [
    # Recursos próprios
    "RP - Recursos Próprios",
    "ICMS",
    "ISS",
    "IRRF",
    "ROYALTIES",
    "COSIP",
    # Federais educação
    "FUNDEB 70%",
    "FUNDEB 30%",
    "FME - Fundo Mun. de Educação",
    "FNDE - Salário-Educação",
    "PNAE - Alimentação Escolar",
    "PNATE - Transporte Escolar",
    "PDDE - Dinheiro Direto na Escola",
    # Federais saúde
    "FMS - Fundo Mun. de Saúde",
    "PAB - Piso de Atenção Básica",
    "PNAB",
    "MAC - Média e Alta Complexidade",
    "FUNPEQ",
    # Federais assistência
    "FMAS - Fundo Mun. de Assistência",
    "FNAS - Fundo Nacional de Assistência",
    "IGDBF / IGDSUAS",
    "PSB - Proteção Social Básica",
    "GESTÃO SUAS",
    "CRIANÇA FELIZ",
    "RENDA CIDADÃ",
    # Convênios e emendas
    "EMENDA PARLAMENTAR",
    "EMENDA ESTADUAL",
    "CONVÊNIO - Estado",
    "CONVÊNIO - União",
    # Outros
    "FPM - Fundo de Participação",
    "INCRA",
    "OUTROS",
]

CATEGORIAS = [
    "Aquisição de material",
    "Serviços de terceiros - PF",
    "Serviços de terceiros - PJ",
    "Folha de pagamento",
    "Diárias e passagens",
    "Combustível",
    "Energia / Água / Telefone",
    "Convênio / Repasse",
    "Obra / Reforma",
    "Tributos e encargos",
    "Aluguel",
    "Outros",
]

TIPOS_PAGAMENTO = [
    "Transferência bancária",
    "Cheque",
    "PIX",
    "Boleto",
    "Repasse interno",
    "Empenho",
    "Outros",
]

# Fornecedores: deixar vazio para a prefeitura ir cadastrando.
# Cada fornecedor cadastrado uma vez na aba "Fornecedores" pode ser reutilizado.
FORNECEDORES_INICIAIS: list[tuple[str, str, str]] = [
    # (Nome usado, Razão Social completa, CNPJ)
    # Deixar vazio — o usuário preenche conforme paga.
]

# -------------------------------------------------------------------
# Estilos
# -------------------------------------------------------------------
COR_HEADER       = "1E5B9E"   # azul executive
COR_HEADER_FG    = "FFFFFF"
COR_CAT_HEADER   = "374151"
COR_CAT_HEADER_FG = "F3F4F6"
COR_ZEBRA_A      = "FFFFFF"
COR_ZEBRA_B      = "F3F6FB"
COR_DESTAQUE     = "FEF3C7"   # amarelo claro para campos obrigatórios opcionais
COR_INSTR_BG     = "EFF6FF"   # azul claríssimo
COR_BORDA        = "CBD5E1"

THIN = Side(style="thin", color=COR_BORDA)
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FONT_TITULO = Font(name="Inter", size=14, bold=True, color="1F2937")
FONT_HEADER = Font(name="Inter", size=10, bold=True, color=COR_HEADER_FG)
FONT_CAT_HEADER = Font(name="Inter", size=10, bold=True, color=COR_CAT_HEADER_FG)
FONT_BODY   = Font(name="Inter", size=10, color="111827")
FONT_INSTR  = Font(name="Inter", size=11, color="1F2937")
FONT_HINT   = Font(name="Inter", size=9, italic=True, color="6B7280")


# -------------------------------------------------------------------
# Helpers
# -------------------------------------------------------------------
def _set_widths(ws, widths: dict[str, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _style_header_row(ws, row: int, max_col: int,
                       fill: str = COR_HEADER, font: Font = FONT_HEADER) -> None:
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_ALL
    ws.row_dimensions[row].height = 34


def _list_validation_from_range(ws, sheet_ref: str, start_row: int, end_row: int) -> DataValidation:
    """Cria DataValidation que aponta para uma faixa de outra aba."""
    formula = f"={sheet_ref}!$A${start_row}:$A${end_row}"
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "Selecione um valor da lista. Para incluir um novo, adicione na aba correspondente."
    dv.errorTitle = "Valor inválido"
    dv.showErrorMessage = True
    return dv


# -------------------------------------------------------------------
# Aba 1: Instruções
# -------------------------------------------------------------------
def aba_instrucoes(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Instruções"

    ws["B2"] = "📋 PADRÃO MUNICIPAL DE DESPESAS — PREFEITURA DE MAUÉS"
    ws["B2"].font = Font(name="Inter", size=16, bold=True, color="1E5B9E")
    ws.merge_cells("B2:G2")

    ws["B3"] = "Modelo único para registro mensal de despesas pagas pela Prefeitura."
    ws["B3"].font = Font(name="Inter", size=11, italic=True, color="6B7280")
    ws.merge_cells("B3:G3")

    linhas = [
        ("",                                                                                ""),
        ("COMO PREENCHER",                                                                  ""),
        ("1.", "Vá para a aba **Lançamentos** e adicione UMA linha por pagamento."),
        ("2.", "Use as **listas suspensas** (▼) nas colunas Secretaria, Recurso, Categoria"
                " e Tipo de Pagamento. Não digite livremente."),
        ("3.", "Para o fornecedor: digite o nome **exatamente como aparece no documento**"
                " (NF/recibo). Se for um fornecedor novo, cadastre primeiro na aba"
                " **Fornecedores** com CNPJ e razão social."),
        ("4.", "Data: use formato **dd/mm/aaaa** (ex: 12/01/2026). As colunas **Mês** e"
                " **Ano** são preenchidas automaticamente."),
        ("5.", "Valor: digite só o número, com vírgula decimal (ex: 53049,15). NÃO inclua"
                " R$, ponto de milhar ou texto."),
        ("",                                                                                ""),
        ("REGRAS IMPORTANTES",                                                              ""),
        ("•", "**Não consolidar fornecedores parecidos.** Se 'LM' e 'L M' são empresas"
                " diferentes, mantenha como duas linhas no catálogo de Fornecedores."),
        ("•", "**Não apagar a estrutura das colunas** ou alterar nome das abas — o sistema"
                " usa o nome exato delas pra importar."),
        ("•", "**Preencha do início ao fim do mês** antes de enviar. Lançamentos parciais"
                " atrasam a importação."),
        ("•", "Para inserir um **novo recurso** (fonte), adicione na aba **Recursos** e ele"
                " aparece automaticamente nas próximas linhas."),
        ("",                                                                                ""),
        ("COLUNAS DA ABA LANÇAMENTOS",                                                      ""),
        ("Data",          "Data efetiva do pagamento. Formato: dd/mm/aaaa."),
        ("Mês / Ano",     "Calculados automaticamente a partir da Data — NÃO preencha."),
        ("Secretaria",    "Lista suspensa. Secretaria responsável pelo gasto."),
        ("Fornecedor",    "Nome usado no documento. Se for novo, cadastre na aba Fornecedores."),
        ("Razão Social",  "Calculado a partir do Fornecedor — NÃO preencha."),
        ("CNPJ",          "Calculado a partir do Fornecedor — NÃO preencha."),
        ("Recurso",       "Lista suspensa. Fonte do recurso (FME, FPM, ICMS, etc)."),
        ("Categoria",     "Lista suspensa. Tipo de despesa (Serviços, Material, etc)."),
        ("Tipo de Pgto",  "Lista suspensa. Forma de pagamento (PIX, Boleto, etc)."),
        ("Descrição",     "Descrição livre. Ex: 'NF 003 SEMAS', 'aluguel jan/26'."),
        ("Valor (R$)",    "Valor pago em reais. Use vírgula decimal: 53049,15"),
        ("Conta",         "Opcional. Número/agência da conta corrente."),
        ("Observações",   "Opcional. Qualquer informação extra."),
        ("",                                                                                ""),
        ("DÚVIDAS? Fale com a equipe de TI.",                                               ""),
    ]
    r = 5
    for esq, dir_ in linhas:
        a = ws.cell(row=r, column=2, value=esq)
        b = ws.cell(row=r, column=3, value=dir_)
        a.font = FONT_INSTR
        b.font = FONT_INSTR
        a.alignment = Alignment(vertical="top", wrap_text=True)
        b.alignment = Alignment(vertical="top", wrap_text=True)
        # se é um cabeçalho de seção (linha sem segundo texto e sem ponto)
        if esq and not dir_ and not esq.startswith(("•", "1.", "2.", "3.", "4.", "5.")):
            a.font = Font(name="Inter", size=12, bold=True, color="1E5B9E")
            a.fill = PatternFill("solid", fgColor=COR_INSTR_BG)
            b.fill = PatternFill("solid", fgColor=COR_INSTR_BG)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
            ws.row_dimensions[r].height = 22
        else:
            ws.row_dimensions[r].height = 18
        r += 1

    _set_widths(ws, {"A": 2, "B": 22, "C": 90, "D": 2})


# -------------------------------------------------------------------
# Aba 2: Lançamentos (principal)
# -------------------------------------------------------------------
COLS_LANC = [
    ("A", "Data",         12, "date"),
    ("B", "Mês",          8,  "formula"),
    ("C", "Ano",          7,  "formula"),
    ("D", "Secretaria",   24, "dropdown"),
    ("E", "Fornecedor",   34, "dropdown"),
    ("F", "Razão Social", 36, "formula"),
    ("G", "CNPJ",         18, "formula"),
    ("H", "Recurso",      28, "dropdown"),
    ("I", "Categoria",    22, "dropdown"),
    ("J", "Tipo de Pgto", 18, "dropdown"),
    ("K", "Descrição",    44, "text"),
    ("L", "Valor (R$)",   14, "money"),
    ("M", "Conta",        14, "text"),
    ("N", "Observações",  30, "text"),
]


def aba_lancamentos(wb: Workbook, total_linhas: int = 2000) -> None:
    ws = wb.create_sheet("Lançamentos")
    ws.freeze_panes = "A2"

    # Cabeçalho
    for i, (col, header, _w, _t) in enumerate(COLS_LANC, start=1):
        ws.cell(row=1, column=i, value=header)
    _style_header_row(ws, 1, len(COLS_LANC))

    # Larguras
    _set_widths(ws, {col: w for col, _h, w, _t in COLS_LANC})

    # Fórmulas e formatação por coluna
    for r in range(2, total_linhas + 2):
        # Mês = MONTH(Data)
        ws[f"B{r}"] = f"=IF(A{r}=\"\",\"\",MONTH(A{r}))"
        # Ano = YEAR(Data)
        ws[f"C{r}"] = f"=IF(A{r}=\"\",\"\",YEAR(A{r}))"
        # Razão Social = VLOOKUP(Fornecedor, Fornecedores!A:C, 2, 0)
        ws[f"F{r}"] = f"=IFERROR(VLOOKUP(E{r},Fornecedores!$A:$C,2,FALSE),\"\")"
        # CNPJ = VLOOKUP(Fornecedor, Fornecedores!A:C, 3, 0)
        ws[f"G{r}"] = f"=IFERROR(VLOOKUP(E{r},Fornecedores!$A:$C,3,FALSE),\"\")"

        # Estilo zebra
        fill = PatternFill("solid", fgColor=(COR_ZEBRA_A if r % 2 == 0 else COR_ZEBRA_B))
        for c in range(1, len(COLS_LANC) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = FONT_BODY
            cell.fill = fill
            cell.border = BORDER_ALL
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        # Formatação específica
        ws[f"A{r}"].number_format = "dd/mm/yyyy"
        ws[f"B{r}"].number_format = "0"
        ws[f"C{r}"].number_format = "0"
        ws[f"L{r}"].number_format = '"R$" #,##0.00;[Red]"-R$"#,##0.00'

        # Marca colunas calculadas em cinza claro (visualmente "não preencha")
        for col_calc in ("B", "C", "F", "G"):
            ws[f"{col_calc}{r}"].fill = PatternFill("solid", fgColor="E5E7EB")
            ws[f"{col_calc}{r}"].font = Font(name="Inter", size=10, color="6B7280", italic=True)

    # Data validations
    dv_sec = _list_validation_from_range(ws, "Secretarias",  2, len(SECRETARIAS) + 1)
    dv_rec = _list_validation_from_range(ws, "Recursos",     2, len(RECURSOS) + 1)
    dv_cat = _list_validation_from_range(ws, "Categorias",   2, len(CATEGORIAS) + 1)
    dv_tip = _list_validation_from_range(ws, "Tipos",        2, len(TIPOS_PAGAMENTO) + 1)
    # Fornecedor: lista cresce dinamicamente — referenciamos as 5000 linhas
    dv_forn = DataValidation(
        type="list",
        formula1="=Fornecedores!$A$2:$A$5001",
        allow_blank=True,
    )
    dv_forn.error = ("Fornecedor não cadastrado. Adicione na aba Fornecedores "
                     "(nome, razão social e CNPJ) antes de usar aqui.")
    dv_forn.errorTitle = "Fornecedor não cadastrado"
    dv_forn.showErrorMessage = True

    rng = f"A2:N{total_linhas + 1}"
    ws.add_data_validation(dv_sec); dv_sec.add(f"D2:D{total_linhas+1}")
    ws.add_data_validation(dv_rec); dv_rec.add(f"H2:H{total_linhas+1}")
    ws.add_data_validation(dv_cat); dv_cat.add(f"I2:I{total_linhas+1}")
    ws.add_data_validation(dv_tip); dv_tip.add(f"J2:J{total_linhas+1}")
    ws.add_data_validation(dv_forn); dv_forn.add(f"E2:E{total_linhas+1}")

    # Validação numérica do Valor: > 0
    dv_val = DataValidation(type="decimal", operator="greaterThan", formula1=0,
                            allow_blank=True)
    dv_val.error = "Valor deve ser um número positivo (use vírgula decimal: 53049,15)."
    dv_val.errorTitle = "Valor inválido"
    dv_val.showErrorMessage = True
    ws.add_data_validation(dv_val); dv_val.add(f"L2:L{total_linhas+1}")

    # Validação de data
    dv_data = DataValidation(type="date", operator="between",
                             formula1="DATE(2020,1,1)",
                             formula2="DATE(2050,12,31)",
                             allow_blank=True)
    dv_data.error = "Data inválida. Use o formato dd/mm/aaaa (ex: 12/01/2026)."
    dv_data.errorTitle = "Data inválida"
    dv_data.showErrorMessage = True
    ws.add_data_validation(dv_data); dv_data.add(f"A2:A{total_linhas+1}")


# -------------------------------------------------------------------
# Abas de catálogo
# -------------------------------------------------------------------
def aba_catalogo_simples(wb: Workbook, nome: str, valores: list[str],
                          header_label: str, hint: str = "") -> None:
    ws = wb.create_sheet(nome)
    ws["A1"] = header_label
    _style_header_row(ws, 1, 1, fill=COR_CAT_HEADER, font=FONT_CAT_HEADER)
    _set_widths(ws, {"A": 50})

    for i, v in enumerate(valores, start=2):
        ws.cell(row=i, column=1, value=v).font = FONT_BODY
        ws.cell(row=i, column=1).border = BORDER_ALL

    if hint:
        ws["C1"] = hint
        ws["C1"].font = FONT_HINT


def aba_fornecedores(wb: Workbook) -> None:
    ws = wb.create_sheet("Fornecedores")
    ws.freeze_panes = "A2"

    headers = ["Nome (como usado nos pagamentos)", "Razão Social", "CNPJ", "Observações"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    _style_header_row(ws, 1, len(headers), fill=COR_CAT_HEADER, font=FONT_CAT_HEADER)

    _set_widths(ws, {"A": 38, "B": 50, "C": 22, "D": 30})

    # Linhas iniciais com formatação
    for r in range(2, 5002):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = FONT_BODY
            cell.border = BORDER_ALL
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if c == 3:  # CNPJ
                cell.number_format = "00\\.000\\.000/0000\\-00"

    for nome, razao, cnpj in FORNECEDORES_INICIAIS:
        ws.append([nome, razao, cnpj, ""])

    # Hint
    ws["F1"] = ("IMPORTANTE: fornecedores parecidos (ex: 'LM' e 'L M') devem ser "
                "mantidos como linhas SEPARADAS. Não consolide.")
    ws["F1"].font = FONT_HINT


# -------------------------------------------------------------------
# Main
# -------------------------------------------------------------------
def main() -> None:
    wb = Workbook()
    aba_instrucoes(wb)
    aba_lancamentos(wb)
    aba_fornecedores(wb)
    aba_catalogo_simples(wb, "Secretarias", SECRETARIAS,
                          "Secretaria",
                          "Para adicionar nova secretaria, inclua aqui.")
    aba_catalogo_simples(wb, "Recursos", RECURSOS,
                          "Recurso / Fonte de Financiamento",
                          "Lista padronizada — não duplique entradas.")
    aba_catalogo_simples(wb, "Categorias", CATEGORIAS,
                          "Categoria de Despesa",
                          "Tipo do gasto (Serviços, Material, etc).")
    aba_catalogo_simples(wb, "Tipos", TIPOS_PAGAMENTO,
                          "Tipo de Pagamento",
                          "Forma de pagamento (PIX, Boleto, etc).")

    # Esconde as abas de catálogo da navegação (mas continuam acessíveis)
    for nome in ("Secretarias", "Recursos", "Categorias", "Tipos"):
        wb[nome].sheet_state = "hidden"

    # Define a aba ativa como Instruções
    wb.active = 0

    wb.save(OUT_FILE)
    print(f"OK — Template gerado em:\n  {OUT_FILE}")
    print(f"  Abas: {wb.sheetnames}")


if __name__ == "__main__":
    main()
